from math import sqrt
import pickle
from dateutil.relativedelta import relativedelta
import tensorflow as tf
import os
# from datetime import datetime

from tensorflow.python.ops.gen_math_ops import rsqrt
from market_data import market_data

from tf_utils import get_strategy, CustomSchedule, create_logger

print(tf.__version__)
import transformers
print(transformers.__version__)
from transformers import TFRobertaForSequenceClassification, RobertaConfig
import pandas as pd

d_model = 32
dummy_input = (
  tf.random.uniform([8, 8], 0, 100, dtype=tf.int32),
  tf.random.uniform([8], 0.5, 1.5, dtype=tf.float32)
  )
from stock_data import stock_data
from utils import time_range
# from stock_model import Transformer

# def to_matrix(l, n):
#   return [l[i:i+n] for i in range(0, len(l), n)]

class cloud_path_initialilzer():
  def __init__(self, path_to_model, cloud=''):
    if cloud  == 'kaggle':
      self.path_to_dataset = '/kaggle/input/securities-dataset/'
      self.path_to_saved_model = '/kaggle/input/securities-transf-saved-model/' #+ path_to_model
      self.output_path = '/kaggle/working/'# + path_to_model
      self.path_to_market_data = '/kaggle/input/market-data/'
      self.path_to_validation = '/kaggle/input/securities-transf-validation/' + path_to_model
      #path_to_embedding_model = '/kaggle/input/poetry-supervised-embedding-model/'
    else:
      path = ''
      # if cloud == 'colab':
      #   path = '/content/drive/My_Drive/Colab_Notebooks/Accounting/'
        # if device == 'TPU':
        #   assert tf.__version__[:3] == '2.2'
      # config = str(d_model) + '/' #'128-512/'
      self.path_to_dataset = "../download_from_CC/result" #path + 'data/'
      self.path_to_saved_model = path + 'saved/' + path_to_model
      self.output_path = path + 'saved/' + path_to_model
      self.path_to_market_data = path + '../data/'
      self.path_to_validation = path + 'saved/' + path_to_model
    from pathlib import Path
    Path(self.output_path).mkdir(parents=True, exist_ok=True)
      #path_to_embedding_model = path + '../../embeddings/saved/' + config

class test_securities(cloud_path_initialilzer):
  pretrained_model_name = 'roberta-base'
  def __init__(self, test_range, train_range, cloud=''):
    super().__init__(str(train_range.start) + '/', cloud)
    self.train_range, self.test_range = train_range, test_range

  def load_data(self, path_test_with_ret):
    from utils import check_file
    if check_file(path_test_with_ret):
      strategy, dev = get_strategy()
      assert dev != 'TPU', "Predict method not working with TPU - https://github.com/huggingface/transformers/issues/12202"
      self.init_dataset(dev)
      if dev != 'CPU':
        self.dataset.nsteps = 64
        # self.dataset.set_batch_sizes_based_on(128)
      with strategy.scope():
        self.load_model()
        with open(path_test_with_ret, 'w') as f:
          from time import time
          start = time()
          for sentence in self.dataset.load_batch(self.test_range, verbose=1, roundup=False):
            data = self.dataset.dataset_from_target(sentence[:1])
            data = strategy.experimental_distribute_dataset(data)
            prediction = self.model.predict(data, verbose=0).logits # not working with TPU
            for s in zip(*sentence, prediction):
              f.write(str(s[4]) + '\t' + s[3] + '\t' + str(s[2]) + '\t' + str(s[1]) + '\t' + str(s[5][0]) + '\n')

            # for inp in data:
            #   # prediction = step_fn(inp)
            #   prediction = strategy.run(step_fn, args=(inp,)).logits.numpy()
            #   for s in zip(*sentence, prediction):
            #     f.write(str(s[4]) + '\t' + s[3] + '\t' + str(s[2]) + '\t' + str(s[1]) + '\t' + str(s[5][0]) + '\n')
            print(time() - start)
    df = pd.read_csv(
      path_test_with_ret, '\t', 
      names=['articleID', 'comp_short', 'publ_time', 'score', 'prediction'],
      parse_dates=[2]
      )
    df['publ_date'] = df.publ_time.dt.date
    df = df.groupby(['articleID', 'comp_short', 'publ_date']).mean()
    return df

  def load_market_data(self):
    import pandas as pd
    file_name = self.path_to_market_data + 'market_data.parquet'
    if not os.path.isfile(file_name):
      print('No market_data.parquet found !!! Create on CPU')
      md = pd.read_excel(self.path_to_market_data + 'market_data.xlsx', engine='openpyxl')
      md = md[pd.notnull(md["A"])]
      md.Date = md.Date.dt.date
      md.set_index('Date', inplace=True)
      md.to_parquet(file_name, compression='gzip')
    else:
      md = pd.read_parquet(file_name)
    return md

  def test_strategy(self, holding_period, sample_size=1, short_pos_adj=1, lag=0):
    # for name, group in grouped:
    #   group = group.groupby('comp_short').agg('mean')
    # day_means = [group.groupby('comp_short').agg('mean') for name, group in grouped]
    # rets_full = [group.iloc[group.ret.argmax()] for name, group in day_means]
    def get_abs_ret(ret, adj):
      # ret = ret['ret']
      if ret < 1:
        return adj * (1 - ret)
      else:
        return ret - 1
    import numpy as np
    self.df['abs_ret'] = np.vectorize(get_abs_ret)(self.df.prediction, short_pos_adj)
    inds = self.df['abs_ret'].groupby('publ_date', group_keys=False).nlargest(sample_size).index
    rets_full = self.df.loc[inds]
    # rets_full = df.loc[df.groupby('publ_date')['abs_ret'].idxmax()]
    rets_full = [(d, c, r.prediction, r.abs_ret) for (d, c), r in rets_full.iterrows()]

    res = []
    for i in range(sample_size):
      rets = rets_full[i::sample_size]
      cash = self.test_sample(rets, holding_period, lag)
      res += [cash]
    return sum(res) / len(res)

  def test_sample(self, rets, holding_period, lag):
    rets += [(rets[-1][0] + relativedelta(days=1), 'A', 1., 0.)]
    quotes = [i[0] in self.md.index for i in rets]
    best = None
    hp_counter = 1
    last = len(rets) - lag
    cash = 100000
    invested = {'comp': 'A', 'quantity': 0, 'ret': 0, 'quote': 1}
    for prev, next, q in zip(rets[:last], rets[lag:], quotes[lag:]):
      hp_counter -= 1
      if best is None or prev[3] > best[3]:
        best = prev
      if not q:
        continue
      try:
        ret = self.df.loc[next[0], invested['comp']].abs_ret
        invested['ret'] = ret
      except KeyError:
        pass
      if best[3] > invested['ret']:
        hp_counter = 0
      if hp_counter <= 0:
        quote = self.md.loc[next[0], invested['comp']]
        cash += (quote - invested['quote']) * invested['quantity']
        with open(self.output_path + '../output.txt', 'a') as f:
          f.write(str(next[0]) + '\t' + str(cash) + '\t' + str(invested['comp']) + '\t' + str(invested['quantity']) + '\n')
        quote = self.md.loc[next[0], best[1]]
        quantity = cash // quote
        if best[2] < 1:
          quantity = -quantity
        invested = {'comp': best[1], 'quantity': quantity, 'ret':best[3], 'quote': quote}
        hp_counter = holding_period
        best = None
    return cash

  def test(self):
    if os.path.exists(self.path_to_saved_model + 'saved.h5'):
      path_test_with_ret = self.output_path + 'test_return_forecast.txt'
      self.md = self.load_market_data()
      df = self.load_data(path_test_with_ret)
      self.df = df.groupby(['publ_date', 'comp_short']).mean()

      # a = test_sample(rets_full, 3, 0)
      a = self.test_strategy(1, 1, 1)
    else:
      print(self.path_to_saved_model + 'saved.h5', 'not found')
            
  def init_model(self):
    config = RobertaConfig.from_pretrained(self.pretrained_model_name)
    config.num_labels = 1
    config.vocab_size = len(self.dataset.tok.vocab)
    # config.hidden_size = 512
    # config.intermediate_size = 2048
    # config.num_attention_heads = 8
    self.model = TFRobertaForSequenceClassification(config)
    # self.model.layers[0].trainable = False

  def load_model(self):
    self.init_model()
    self.model(dummy_input[0])
    self.model.load_weights(self.path_to_saved_model + 'saved.h5')

  def init_dataset(self, dev='CPU', nsteps=10):
    # train_range, test_range = self.set_date_ranges()
    self.dataset = stock_data(
      self, nsteps, #self.train_range, self.test_range,
      device=dev, 
      # model_name=self.pretrained_model_name
      )

  # def get_train_range(self, test_range):
  #   # test_range = time_range(datetime(2021, 1, 14), datetime(2021, 1, 15))
  #   train_range = time_range(test_range.end, test_range.end + relativedelta(months=1))
  #   # train_range = time_range(last_date - relativedelta(months=1) + relativedelta(days=1), last_date)
  #   return train_range#, test_range

class stock(test_securities):
  def load_embeddings(self):
    pass

  def train(self, fine_tune=False):
    import time
    strategy, dev = get_strategy()
    assert dev != 'GPU', 'train on TPU or CPU !!!'
    nsteps = 1
    self.init_dataset(dev, nsteps)
    self.dataset.prepare_for_training()
    eval_batch_size = tf.cast(self.dataset.validation_data._batch_size, tf.float32)
    self.dataset.validation_data = strategy.experimental_distribute_dataset(self.dataset.validation_data)

    # logger = create_logger(self.output_path + '../')

    trainParameters={}
    import json

    @tf.function()
    def step_fn(input, labels):
      #print('retracing')
      with tf.GradientTape() as tape:
        output = self.model(input, training=True)
        loss = tf.keras.losses.MSE(
            labels, output.logits)
        loss = tf.nn.compute_average_loss(loss, global_batch_size=batch_size)
      grads = tape.gradient(loss, self.model.trainable_variables)
      optimizer.apply_gradients(list(zip(grads, self.model.trainable_variables)))
      training_loss.update_state(loss * strategy.num_replicas_in_sync)

    @tf.function()
    def eval_step_fn(input, labels):
      output = self.model(input)
      loss = tf.keras.losses.MSE(
          labels, output.logits)
      loss = tf.nn.compute_average_loss(loss, global_batch_size=eval_batch_size)
      eval_loss.update_state(loss * strategy.num_replicas_in_sync)
      # eval_accuracy.update_state(labels, logits)

    with strategy.scope():
      self.init_model()

      def scheduler(epoch, lr):
        return rsqrt(epoch) * 1e-4

      class RsqrtSchedule(tf.keras.optimizers.schedules.LearningRateSchedule):
        def __call__(self, step):
          init_rate = 0.0003
          rate = tf.math.rsqrt(step + 1) * init_rate
          # flat_steps = 1
          # rate = tf.cond(step < flat_steps, lambda: init_rate, lambda: tf.math.rsqrt(step - flat_steps + 1) * init_rate)
          # mult = 4 / 1666 * step + 1
          # rate /= mult
          return rate

      # learning_rate = CustomSchedule(768, 17000)
      learning_rate = RsqrtSchedule()
      # learning_rate = tf.keras.optimizers.schedules.ExponentialDecay(
      #   3e-4,
      #   decay_steps=1,
      #   decay_rate=0.997,
      #   )
      #lr = learning_rate(259.)
      from transformers.optimization_tf import AdamWeightDecay
      optimizer = AdamWeightDecay(learning_rate, epsilon=1e-6, weight_decay_rate=0.01)
      optimizer._HAS_AGGREGATE_GRAD = False
      self.model.compile(
          loss=tf.keras.losses.MeanSquaredError(reduction=tf.keras.losses.Reduction.NONE),
          optimizer=optimizer,
      )
      training_loss = tf.keras.metrics.Mean('training_loss', dtype=tf.float32)
      eval_loss = tf.keras.metrics.Mean('eval_loss', dtype=tf.float32)
      # eval_accuracy = tf.keras.metrics.SparseCategoricalAccuracy(
      #   'eval_accuracy', dtype=tf.float32)
      # model.layers[-1].encoder.trainable = False
      import pickle
      if not os.path.isfile(self.path_to_saved_model + 'trainParameters.txt'):
        print('creating model...')
        trainParameters={'bestLoss':float('inf'), 'iter':0}
        self.model(dummy_input[0])
        self.load_embeddings()
        # self.model.summary()
        # c = self.model.roberta.embeddings.count_params()
        # c = self.model.roberta.encoder.count_params()
      else:
        print('loading model...')
        trainParameters = json.load(open(self.path_to_saved_model + "trainParameters.txt"))
        # hist = model.fit(dummy_input, dummy_input, verbose=0)
        batch_size = 8
        if fine_tune:
          # step_fn(*dummy_input)
          self.model.roberta.trainable = True
          strategy.run(step_fn, args=(dummy_input[0], dummy_input[1]))
          self.model.load_weights(self.path_to_saved_model + 'saved.h5')
          optimizer.iterations.assign(trainParameters['iter'])
          # self.model.summary()
        else:
          strategy.run(step_fn, args=(dummy_input[0], dummy_input[1]))
          self.model.load_weights(self.path_to_saved_model + 'saved.h5')
          optimizer.iterations.assign(trainParameters['iter'])
          with open(self.path_to_saved_model + 'optimizer.pkl', 'rb') as f:
              weight_values = pickle.load(f)
          optimizer.set_weights(weight_values)

    firstIter = trainParameters['iter']
    start = time.time()
    print('training started...')
    min_tr_loss = 0.007
    min_eval_loss = 0.00091
    for _ in range(1):
      for i, data in zip(
        range(firstIter, 1000000, nsteps), 
        self.dataset.next(verbose=1, roundup=True)
        ):
        if i > 176:
          return
        batch_size = tf.cast(data._batch_size, tf.float32)
        data = strategy.experimental_distribute_dataset(data)
        training_loss.reset_states()
        for input, labels in data:
          # print('step')
          # step_fn(input, labels)
          strategy.run(step_fn, args=(input, labels))
          tl = training_loss.result()
        # break
          # out = 'Loss {:.6f} r: {:.6f}'.format(
          #     round(float(tl), 6),
          #     optimizer._decayed_lr(tf.float32),
          #     )
          # print(out)
          if i > 45 and min_tr_loss > tl:
            min_tr_loss = tl
          # if True:#(i + nsteps) % 1000 == 0:
            eval_loss.reset_states()
            # batch_size = eval_batch_size
            # eval_start = time.time()
            for input, labels in self.dataset.validation_data: 
              # batch_size = tf.cast(tf.shape(labels)[0], tf.float32)
              # eval_step_fn(input, labels)
              strategy.run(eval_step_fn, args=(input, labels))
            # print("evaluation time: ", time.time() - eval_start)
            evaluation_loss = float(eval_loss.result())
            r = optimizer._decayed_lr(tf.float32)
            out = 'Date\t{}\tEpoch\t{}\tLoss\t{:.6f}\tVal loss\t{:.8f}\ttime:\t{:.0f}\tlr:\t{:.6f}\n'.format(
              str(self.train_range.start),
              i + nsteps,
              round(float(tl), 6),
              round(evaluation_loss, 6),
              time.time() - start,
              r,
              )
            print(out)
            # logger.warning(out)
            if evaluation_loss < min_eval_loss:
              self.model.save_weights(self.output_path + 'saved.h5')
              with open(self.output_path + '../file.log', 'a') as f:
                f.write(out)
              min_eval_loss = evaluation_loss
              # return
            pass

            #   if evaluation_loss <= trainParameters['bestLoss']:
            #       print('saving...')
            #       self.model.save_weights(self.output_path + 'saved.h5')
            #       weight_values = optimizer.get_weights()
            #       with open(self.output_path + 'optimizer.pkl', 'wb') as f:
            #           pickle.dump(weight_values, f)
            #       trainParameters['bestLoss'] = evaluation_loss
            #       trainParameters['iter'] = i + nsteps
            #       json.dump(trainParameters, open(self.output_path + "trainParameters.txt",'w'))

  def convert_text_to_indices(self, sentence):
    sentence = self.dataset.bpe.process_line(sentence)
    enc_input = [self.dataset.get_word_index(word) for word in sentence.split()] #+ [self.dataset.PAD_token] * 4
    # dec_input = [self.dataset.SOS_token] #+ enc_input
    enc_input = tf.convert_to_tensor(enc_input)
    input_tensor = tf.reshape(enc_input,(1,-1))
    return input_tensor

  def predict(self, model, inp):
    output = model(inp)
    cat = tf.argmax(output, -1).numpy().squeeze((-1))[0]
    account = self.dataset.index2word[cat]
    return account

  def save_evaluation_results(self):
    self.load_model()
    self.dataset.prepare_for_training('accounting.pkl')
    freq = [0] * len(self.dataset.index2word_tar)
    for v in self.dataset.buckets_files.values():
      if v.size != 0:
        for l in v[:, -1].tolist():
          freq[l] += 1
    import openpyxl
    name = 'analysis_validation.xlsx' #'Карточка счета 51 за January 2021 Общество с ограниченной ответственностью  Трейд Хаус Компани .xlsx'
    wb = openpyxl.load_workbook(name)
    ws = wb['Sheet2']
    output = self.model.predict(self.dataset.validation_data)[0]
    cat = tf.argmax(output, -1).numpy()
    probs = tf.nn.softmax(output)
    for ind, (inp, labels) in enumerate(self.dataset.validation_data.unbatch().batch(1)):
      best = cat[ind]
      account = self.dataset.index2word_tar[best]
      sent = []
      for i in inp['input_ids'][0].numpy().tolist():
        if i == self.dataset.PAD_token:
          break
        s = self.dataset.index2word[i]
        sent += [s]
      ws['A' + str(ind + 2)] = ' '.join(sent)
      ws['B' + str(ind + 2)] = account
      ws['C' + str(ind + 2)] = self.dataset.index2word_tar[labels[0].numpy()]
      ws['D' + str(ind + 2)] = probs[ind][best].numpy()
    # for ind, f in enumerate(freq):
    #   ws['G' + str(ind + 2)] = self.dataset.index2word_tar[ind]
    #   ws['H' + str(ind + 2)] = f
      
    wb.save(name)

from transformers import TFDistilBertForSequenceClassification, DistilBertConfig, DistilBertTokenizerFast

class stock_distilled_bert(stock):
  pretrained_model_name = 'distilbert-base-uncased-finetuned-sst-2-english'
  def init_model(self):
    config = DistilBertConfig.from_pretrained(self.pretrained_model_name)
    config.num_labels = 1
    # config.vocab_size = len(self.dataset.tok.vocab)
    # config.hidden_size = 512
    # config.intermediate_size = 2048
    # config.num_attention_heads = 8
    self.model = TFDistilBertForSequenceClassification(config)
    self.model.layers[0].trainable = False

  def init_dataset(self, dev='CPU', nsteps=10):
    # train_range, test_range = self.set_date_ranges()
    self.dataset = stock_data(
      self, nsteps, #self.train_range, self.test_range,
      tokenizer=DistilBertTokenizerFast.from_pretrained(self.pretrained_model_name),
      device=dev, #enc='cp1251', min_frequency=43, 
      )

  def load_embeddings(self):
    pretrained = TFDistilBertForSequenceClassification.from_pretrained(self.pretrained_model_name)
    self.model.layers[0].set_weights(pretrained.layers[0].get_weights())
    # self.model.summary()
    
from transformers import TFConvBertForSequenceClassification, ConvBertConfig, ConvBertTokenizerFast

class stock_convbert(stock):
  def init_model(self):
    config = ConvBertConfig()
    config.num_labels = 1
    config.vocab_size = len(self.dataset.tok.vocab)
    # config.hidden_size = 1024
    # config.embedding_size = 1024
    # config.intermediate_size = 4096
    # config.num_attention_heads = 16
    # config.num_hidden_layers = 24
    self.model = TFConvBertForSequenceClassification(config)

  def init_dataset(self, dev='CPU', nsteps=10):
    super().init_dataset(dev, nsteps)
    if dev != 'CPU':
      self.dataset.set_batch_sizes_based_on(64)

class stock_convbert_from_pretrained(stock_convbert):
  pretrained_model_name = 'YituTech/conv-bert-base'
  def init_model(self):
    config = ConvBertConfig.from_pretrained(self.pretrained_model_name)
    config.num_labels = 1
    self.model = TFConvBertForSequenceClassification(config)
    self.model.layers[0].trainable = False

  def init_dataset(self, dev='CPU', nsteps=10):
    # train_range, test_range = self.set_date_ranges()
    self.dataset = stock_data(
      self, nsteps, #self.train_range, self.test_range,
      tokenizer=ConvBertTokenizerFast.from_pretrained(self.pretrained_model_name),
      device=dev, #enc='cp1251', min_frequency=43, 
      )

  def load_embeddings(self):
    pretrained = TFConvBertForSequenceClassification.from_pretrained(self.pretrained_model_name)
    self.model.layers[0].set_weights(pretrained.layers[0].get_weights())
    # self.model.summary()
    
from transformers import TFLongformerForSequenceClassification, LongformerConfig

class stock_longformer(stock):
  def init_model(self):
    config = LongformerConfig()
    config.num_labels = 1
    config.vocab_size = len(self.dataset.tok.vocab)
    self.model = TFLongformerForSequenceClassification(config)

  # def init_dataset(self, dev='CPU', nsteps=10):
  #   train_range = time_range(datetime(2010, 1, 1), datetime(2021, 1, 1))
  #   test_range = time_range(datetime(2021, 1, 1), datetime(2021, 1, 10))
  #   self.dataset = stock_data(
  #     self.path_to_dataset, nsteps, train_range, test_range,
  #     device=dev, enc='cp1251', min_frequency=43, 
  #     )
    # if dev != 'CPU':
    #   self.dataset.set_batch_sizes_based_on(64)

def group_companies():
  md = market_data()
  # c = md.returns.corr()
  # c = md.returns.A.corr(md.returns.K)
  d = md.returns.stack().std()
  m = md.returns.stack().mean()
  # M = len(md.returns.index)
  # N = len(md.returns.columns)
  # import numpy as np
  # df_rand = pd.DataFrame(np.random.normal(m, d, (M, N)), columns=md.returns.columns, index=md.returns.index)
  # md.returns[pd.isnull(md.returns)] = df_rand[pd.isnull(md.returns)]
  md.returns.dropna(axis=0, how='all', inplace=True)
  md.returns.dropna(axis=1, inplace=True)
  data = md.returns.to_numpy().T
  nclusters = 10
  from sklearn.cluster import KMeans
  kmeans = KMeans(n_clusters=nclusters).fit(data)

  comp_article_count_path = 'data/comp_article_count.pkl'
  if os.path.exists(comp_article_count_path):
    with open(comp_article_count_path, 'rb') as f:
      comp_article_count = pickle.load(f)
  else:
    def comps_it():
      from stock_data import stock_data_it
      stock_it = stock_data_it('../data/', "../download_from_CC/result")
      for article in stock_it.sentences_it(time_range(), roundup=False, verbose=1):
        yield article[3]

    from collections import Counter
    comp_article_count = Counter(comps_it())
    with open(comp_article_count_path, 'wb') as f:
      pickle.dump(comp_article_count, f)

  group_count = [0] * nclusters
  comps_grouped = [[] for _ in range(nclusters)]
  for comp, g in zip(md.returns.columns, kmeans.labels_):
    group_count[g] += comp_article_count[comp]
    comps_grouped[g] += [comp]

    # with open('output.txt', 'w') as f:
    #   for k, v in c.items():
    #     f.write(k + '\t' + str(v) + '\n')

  sorted_groups = [(g, s) for g, s in enumerate(group_count)]
  sorted_groups.sort(key=lambda tup: tup[1], reverse=True)
  num_articles = sum(group_count)
  max_len = num_articles / len(group_count)

  for i, g in enumerate(sorted_groups):
    if g[1] > max_len:
      for comp in comps_grouped[g[0]]:
        for gr in range(i + 1, len(sorted_groups)):
            
          pass

# if __name__ == '__main__':
  # from pathlib import Path
  # import pandas as pd
  # pathlist = Path("../download_from_CC/result").glob('*.parquet')
  # # comps = []
  # for i, path in enumerate(sorted(pathlist)):
  #   # if i > 40:
  #   #   break
  #   table = pd.read_parquet(path)
  #   comps += table.comp_short.unique().tolist()
  # for c in comps:
  #   print(c)

path_to_working_days = ''
if 'kaggle' in globals():
  cloud = 'kaggle'
  path_to_working_days = '/kaggle/input/working-days/'
elif 'colab' in globals():
  cloud = 'colab'
else:
  cloud = ''
df = pd.read_excel(
  path_to_working_days + 'analysis_net_of_index_best_eval_loss.xlsx', 
  sheet_name='work days sample', 
  header=0,
  engine='openpyxl'
  )
i = 83
j = i + 1
for eval_start, train_start, test_start, test_end in zip(df.eval_start.dt.date[i:j], df.train_start.dt.date[i:j], df.test_start.dt.date[i:j], df.test_end.dt.date[i:j]):
  print('!!!!!! processing training period started on ', train_start)
  s = stock_convbert(time_range(eval_start, train_start), time_range(train_start, test_start), cloud)
  # s.init_dataset()
  # s.dataset.prepare_for_training()
  s.train()
  # stock_convbert(time_range(test_start), time_range(train_start, test_start), cloud).test()
  pass
pass
